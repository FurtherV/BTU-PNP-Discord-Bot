from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .dates import german_date
from .models import Registration, Survey


def _safe(value: str) -> str:
    if value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _role(registration: Registration) -> str:
    if registration.is_player and registration.is_dm:
        return "Spieler & DM"
    return "Spieler" if registration.is_player else "DM"


def create_workbook(survey: Survey, registrations: list[Registration]) -> BytesIO:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "Übersicht"
    dates = tuple(survey.week_start.fromordinal(survey.week_start.toordinal() + i) for i in range(7))
    overview.append(["Discord-ID", "Discord-Name", "Rolle", *[german_date(day) for day in dates], "Anmerkung"])
    for registration in registrations:
        available = set(registration.available_dates)
        overview.append([
            str(registration.user_id), _safe(registration.display_name), _role(registration),
            *[("X" if day in available else "") for day in dates], _safe(registration.notes),
        ])

    details = workbook.create_sheet("Einzelantworten")
    details.append([
        "Discord-ID", "Discord-Name", "Spieler", "DM", "Verfügbare Tage",
        "Anmerkung", "Erstellt", "Geändert",
    ])
    for registration in registrations:
        details.append([
            str(registration.user_id), _safe(registration.display_name),
            "Ja" if registration.is_player else "Nein", "Ja" if registration.is_dm else "Nein",
            ", ".join(german_date(day) for day in registration.available_dates),
            _safe(registration.notes), registration.created_at.replace(tzinfo=None),
            registration.updated_at.replace(tzinfo=None),
        ])

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="5865F2")
            cell.alignment = Alignment(wrap_text=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in range(1, sheet.max_column + 1):
            values = [str(sheet.cell(row, column).value or "") for row in range(1, sheet.max_row + 1)]
            sheet.column_dimensions[get_column_letter(column)].width = min(max(map(len, values)) + 2, 45)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
