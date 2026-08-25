from datetime import date, datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from pnp_bot.export import create_workbook
from pnp_bot.models import Registration, Survey


def test_workbook_contains_matrix_details_and_escapes_formulas() -> None:
    tz = ZoneInfo("Europe/Berlin")
    now = datetime(2026, 8, 2, tzinfo=tz)
    survey = Survey(
        id=1, guild_id=2, year=2026, month=8, channel_id=3, message_id=4,
        week_start=date(2026, 8, 24), week_end=date(2026, 8, 30),
        deadline=datetime(2026, 8, 17, 23, 59, tzinfo=tz), state="open",
        created_at=now, closed_at=None, reminder_last_run_at=None, reminder_completed_at=None,
    )
    registration = Registration(
        id=1, survey_id=1, user_id=42, display_name="=FORMULA", is_player=True,
        is_dm=True, notes="+gefährlich", available_dates=(date(2026, 8, 24),),
        created_at=now, updated_at=now,
    )
    output = create_workbook(survey, [registration])
    workbook = load_workbook(BytesIO(output.getvalue()))
    assert workbook.sheetnames == ["Übersicht", "Einzelantworten"]
    assert workbook["Übersicht"]["B2"].value == "'=FORMULA"
    assert workbook["Übersicht"]["D2"].value == "X"
    assert workbook["Einzelantworten"]["F2"].value == "'+gefährlich"
